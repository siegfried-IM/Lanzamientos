"""
build_dashboard.py
Genera dashboard.html (autocontenido) con dos segmentos:
  1) IQVIA — Pack Siegfried vs Pack Total Mercado + MS%, comparable por Molécula o por ATC.
  2) Venta Interna (QLICK) — packs por Familia/Presentación.

Lee los 3 Excel del mismo directorio usando openpyxl read_only
(para evitar el lock que aparece con OneDrive en algunas situaciones).
"""

import json
import re
import os
import sys
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

# Forzar stdout a UTF-8 en Windows (la consola cp1252 falla con flechas/acentos)
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

HERE = Path(__file__).parent
F_MAESTRO = HERE / "MAESTRO.xlsx"
F_IQVIA   = HERE / "IQUVIA_PM.xlsx"
F_QLICK   = HERE / "QLICK_VTA_INTERNA.xlsx"
OUT_HTML  = HERE / "dashboard.html"

ES_MONTHS = {
    "Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Ago": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dic": 12,
}
EN_MONTHS = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}
ES_LABEL = {v: k for k, v in ES_MONTHS.items()}


def read_sheet(path: Path, sheet=None):
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def clean_header(s):
    if s is None:
        return ""
    return str(s).replace("\n", " ").strip()


def parse_iqvia_month(label: str):
    # "May 2021 Units" -> (2021, 5)
    m = re.match(r"([A-Za-z]{3,9})\s+(\d{4})\s+Units", label)
    if not m:
        return None
    mon_str, year = m.group(1)[:3], int(m.group(2))
    mon = EN_MONTHS.get(mon_str)
    if mon is None:
        return None
    return (year, mon)


def parse_qlick_month(label: str):
    # "Mar-2024" -> (2024, 3)
    m = re.match(r"([A-Za-zñÑ]+)-(\d{4})", label or "")
    if not m:
        return None
    mon_str, year = m.group(1)[:3], int(m.group(2))
    mon = ES_MONTHS.get(mon_str.capitalize())
    if mon is None:
        return None
    return (year, mon)


def yyyymm_label_es(year, month):
    return f"{ES_LABEL[month]}-{year}"


def to_num(x):
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s in ("", "-", "—", "NaN", "nan"):
        return 0.0
    s = s.replace(".", "").replace(",", ".") if s.count(",") == 1 and s.count(".") > 1 else s
    try:
        return float(s)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# 1. Cargar IQVIA
# ---------------------------------------------------------------------------
print("→ Cargando IQVIA…")
iqvia_rows = read_sheet(F_IQVIA)
iqvia_header = [clean_header(c) for c in iqvia_rows[0]]
iqvia_data = iqvia_rows[1:]

col = {name: i for i, name in enumerate(iqvia_header)}
i_manuf = col["Manufacturer"]
i_util  = col["Utilizar"]
i_mol   = col["Molecules Short"]
i_atc   = col["ATC-4"]
i_prod  = col["Product"]
i_pack  = col["Pack"]

# Identificar índices de columnas mensuales en IQVIA
iqvia_month_idx = []     # list of (year, month, col_idx)
for i, h in enumerate(iqvia_header):
    ym = parse_iqvia_month(h)
    if ym:
        iqvia_month_idx.append((ym[0], ym[1], i))
iqvia_month_idx.sort()
print(f"  meses IQVIA: {len(iqvia_month_idx)} ({iqvia_month_idx[0][:2]} → {iqvia_month_idx[-1][:2]})")

# Filtrar filas reales (saltar Grand Total y filas vacías)
def is_data_row(r):
    manuf = r[i_manuf]
    return manuf is not None and str(manuf).strip().lower() not in ("", "grand total")

iqvia_real = [r for r in iqvia_data if is_data_row(r)]
print(f"  filas IQVIA con datos: {len(iqvia_real)}")

def is_utilizar_si(v):
    return v is not None and str(v).strip().upper() == "SI"

siegfried_si = [r for r in iqvia_real
                if str(r[i_manuf]).strip().upper() == "SIEGFRIED"
                and is_utilizar_si(r[i_util])]
print(f"  filas Siegfried con Utilizar=SI: {len(siegfried_si)}")

# Mapeo Producto (SIE) → Familia (alineado a Gran Familia de QLICK)
PRODUCT_TO_FAMILY = {
    "ACNECLIN PBA (SIE)":  "ACNECLIN",
    "GELACNE (SIE)":       "ACNECLIN",
    "ALIDIAL L (SIE)":     "ALIDIAL",
    "ALERGICAL (SIE)":     "ALIDIAL",
    "BREXIL (SIE)":        "BREXIL",
    "EMPAX (SIE)":         "EMPAX",
    "EMPAX MET (SIE)":     "EMPAX MET",
    "ISIS NAT (SIE)":      "ISIS",
    "TELPRES (SIE)":       "TELPRES",
    "TRIP +45 (SIE)":      "TRIP",
    "TRIP D3 PLUS (SIE)":  "TRIP",
    "TRIP MAGNESIO (SIE)": "TRIP",
    "VALQUIR (SIE)":       "VALQUIR",
}

# Nombres "lindos" de familias para mostrar en el dashboard
# (las claves internas siguen siendo las del PRODUCT_TO_FAMILY arriba).
FAMILY_DISPLAY_NAMES = {
    "ALIDIAL": "Alidial L",
    "ISIS":    "Isis Nat",
}
def display_family(name):
    return FAMILY_DISPLAY_NAMES.get((name or "").upper(), name)

# Molécula → Familia (vía las filas Siegfried Utilizar=SI)
mol_to_family = {}
for r in siegfried_si:
    fam = PRODUCT_TO_FAMILY.get(str(r[i_prod]).strip())
    mol = str(r[i_mol]).strip() if r[i_mol] else None
    if mol and fam:
        mol_to_family[mol] = fam

atc_to_family = {}
for r in siegfried_si:
    fam = PRODUCT_TO_FAMILY.get(str(r[i_prod]).strip())
    atc = str(r[i_atc]).strip() if r[i_atc] else None
    if atc and fam:
        atc_to_family[atc] = fam

print(f"  moléculas con Siegfried-SI: {len(mol_to_family)}")
print(f"  ATCs con Siegfried-SI:      {len(atc_to_family)}")

# ---------------------------------------------------------------------------
# 2. Determinar la lista COMPLETA de meses disponibles (con ventas Siegfried > 0)
#    Y la ventana DEFAULT = últimos 6 meses con ventas.
# ---------------------------------------------------------------------------
si_totals_by_month = {}
for r in siegfried_si:
    for (y, m, ci) in iqvia_month_idx:
        si_totals_by_month[(y, m)] = si_totals_by_month.get((y, m), 0.0) + to_num(r[ci])

DEFAULT_WINDOW_SIZE  = 6
MIN_PRODUCTS_PER_YEAR = 3   # exigir ≥ N productos Siegfried distintos con ventas para incluir el año

# Cuántos productos Siegfried-SI distintos tuvieron ventas en cada año
products_per_year = {}
for r in siegfried_si:
    prod = str(r[i_prod]).strip()
    for (y, m, ci) in iqvia_month_idx:
        if to_num(r[ci]) > 0:
            products_per_year.setdefault(y, set()).add(prod)
active_years = {y for y, prods in products_per_year.items()
                if len(prods) >= MIN_PRODUCTS_PER_YEAR}
print(f"  productos Siegfried por año: " +
      ", ".join(f"{y}={len(p)}" for y, p in sorted(products_per_year.items())))
print(f"  años retenidos (≥{MIN_PRODUCTS_PER_YEAR} productos): {sorted(active_years)}")

months_with_sales = sorted([ym for ym, v in si_totals_by_month.items()
                            if v > 0 and ym[0] in active_years])
if not months_with_sales:
    raise SystemExit("No hay meses que cumplan el umbral de productos activos")
window_default = months_with_sales[-DEFAULT_WINDOW_SIZE:]
print(f"  meses disponibles tras filtro: {len(months_with_sales)}")
print(f"  ventana default (últimos {DEFAULT_WINDOW_SIZE}): {window_default[0]} → {window_default[-1]}")

# Mapear cada mes disponible al label en español + índices de columna en IQVIA
# all_months = [(label, year, month, iqvia_col_idx)]
all_months = []
for (y, m) in months_with_sales:
    for (yy, mm, ci) in iqvia_month_idx:
        if yy == y and mm == m:
            all_months.append((yyyymm_label_es(y, m), y, m, ci))
            break

# Alias usado por código antiguo más abajo (compatibilidad)
window = window_default
window_iqvia_cols = [ci for (_l, _y, _m, ci) in all_months
                       if (_y, _m) in window]

# ---------------------------------------------------------------------------
# 3. Cargar MAESTRO para mapear nombres IQVIA → QLICK (más legibles)
# ---------------------------------------------------------------------------
maestro_rows = read_sheet(F_MAESTRO)
# Header del MAESTRO está en row index 6
m_header = [clean_header(c) for c in maestro_rows[6]]
m_data   = maestro_rows[7:]
mcol = {name: i for i, name in enumerate(m_header)}

# (NOMBRE_IQUVIA, PRESENTACION_IQUVIA) -> (NOMBRE_QLICK, PRESENTACION_QLICK)
iqvia_to_qlick_names = {}
for r in m_data:
    nq = str(r[mcol["NOMBRE_QLICK"]]        or "").strip()
    pq = str(r[mcol["PRESENTACION_QLICK"]]  or "").strip()
    ni = str(r[mcol["NOMBRE_IQUVIA"]]       or "").strip()
    pi = str(r[mcol["PRESENTACION_IQUVIA"]] or "").strip()
    if ni and pi:
        iqvia_to_qlick_names[(ni, pi)] = (nq, pq)
print(f"  mapeos IQVIA→QLICK en MAESTRO: {len(iqvia_to_qlick_names)}")

# ---------------------------------------------------------------------------
# 4. Agregaciones IQVIA — totales de mercado por Molécula y por ATC
#     + lista de productos Siegfried con sus presentaciones
#    (Por cada mes disponible: dict label_mes → valor)
# ---------------------------------------------------------------------------
month_labels = [t[0] for t in all_months]  # ej. ["Mar-2024", "Abr-2024", ...]

mercado_by_mol = {}   # mol -> {label: total}
mercado_by_atc = {}   # atc -> {label: total}
for r in iqvia_real:
    mol = str(r[i_mol] or "").strip()
    atc = str(r[i_atc] or "").strip()
    for (label, _y, _m, ci) in all_months:
        v = to_num(r[ci])
        if v == 0:
            # Igual lo seteamos para que el dict tenga la clave
            if mol: mercado_by_mol.setdefault(mol, {}).setdefault(label, 0.0)
            if atc: mercado_by_atc.setdefault(atc, {}).setdefault(label, 0.0)
            continue
        if mol:
            d = mercado_by_mol.setdefault(mol, {})
            d[label] = d.get(label, 0.0) + v
        if atc:
            d = mercado_by_atc.setdefault(atc, {})
            d[label] = d.get(label, 0.0) + v

# Construir lista de productos (un bloque por Producto Siegfried)
prods_dict = {}
for r in siegfried_si:
    prod_key = str(r[i_prod]).strip()          # ej. "BREXIL (SIE)"
    pack_iqv = str(r[i_pack] or "").strip()    # ej. "BREXIL TABL RECUBIE 1.00MG x 30"
    mol      = str(r[i_mol]  or "").strip()
    atc      = str(r[i_atc]  or "").strip()
    fam      = display_family(PRODUCT_TO_FAMILY.get(prod_key, ""))

    # Para diferenciar productos (TRIP MAGNESIO vs TRIP D3 PLUS vs TRIP +45)
    # usamos el nombre IQVIA sin " (SIE)" — NOMBRE_QLICK del MAESTRO es genérico.
    prod_display = prod_key.replace(" (SIE)", "").strip()
    _, qlick_pres = iqvia_to_qlick_names.get(
        (prod_key, pack_iqv),
        (prod_display, pack_iqv)
    )

    if prod_key not in prods_dict:
        prods_dict[prod_key] = {
            "producto":  prod_display,
            "familia":   fam,
            "molecula":  mol,
            "atc":       atc,
            "presentaciones": [],
        }
    prods_dict[prod_key]["presentaciones"].append({
        "label":      qlick_pres,
        "iqvia_pack": pack_iqv,
        "si":         {label: to_num(r[ci]) for (label, _y, _m, ci) in all_months},
    })

iqvia_products = []
for prod_key, p in prods_dict.items():
    p["mercado_molecula"] = mercado_by_mol.get(p["molecula"], {l: 0.0 for l in month_labels})
    p["mercado_atc"]      = mercado_by_atc.get(p["atc"],      {l: 0.0 for l in month_labels})
    p["presentaciones"].sort(key=lambda x: x["label"])
    iqvia_products.append(p)

iqvia_products.sort(key=lambda p: (p["familia"], p["producto"]))
print(f"  productos Siegfried (bloques IQVIA): {len(iqvia_products)} "
      f"con {sum(len(p['presentaciones']) for p in iqvia_products)} presentaciones totales")

# ---------------------------------------------------------------------------
# 4. Cargar QLICK (Venta Interna)
# ---------------------------------------------------------------------------
print("→ Cargando QLICK Venta Interna…")
qlick_rows = read_sheet(F_QLICK)
# Header está en row index 1 (la 2da fila)
qlick_header = [clean_header(c) for c in qlick_rows[1]]
qlick_data = qlick_rows[2:]

qcol = {name: i for i, name in enumerate(qlick_header)}
q_soc   = qcol["Sociedad"]
q_gfam  = qcol["Gran Familia"]
q_fam   = qcol["Familia"]
q_prod  = qcol["Producto"]
q_pres  = qcol["Presentación"]
q_code  = qcol["Codigo Producto"]

qlick_month_idx = []
for i, h in enumerate(qlick_header):
    ym = parse_qlick_month(h)
    if ym:
        qlick_month_idx.append((ym[0], ym[1], i))
qlick_month_idx.sort()
print(f"  meses QLICK: {len(qlick_month_idx)} ({qlick_month_idx[0][:2]} → {qlick_month_idx[-1][:2]})")

# Solo SKU reales: Codigo Producto presente y no es "Totales" en columnas jerárquicas
def is_sku_row(r):
    cod = str(r[q_code] or "").strip()
    if not cod:
        return False
    for ci in (q_gfam, q_fam, q_prod):
        v = str(r[ci] or "").strip().lower()
        if v == "totales":
            return False
    return True

qlick_skus = [r for r in qlick_data if is_sku_row(r)]
print(f"  SKUs en QLICK: {len(qlick_skus)}")

# Mapear cada mes disponible (de all_months IQVIA) al índice en QLICK (puede no existir)
qlick_col_by_label = {}
for (label, y, m, _ci) in all_months:
    ci = None
    for (yy, mm, c) in qlick_month_idx:
        if yy == y and mm == m:
            ci = c
            break
    qlick_col_by_label[label] = ci

venta_interna_rows = []
for r in qlick_skus:
    rec = {
        "gran_familia": str(r[q_gfam] or "").strip(),
        "familia":      str(r[q_fam]  or "").strip(),
        "producto":     str(r[q_prod] or "").strip(),
        "presentacion": str(r[q_pres] or "").strip(),
        "codigo":       str(r[q_code] or "").strip(),
        "data":         {label: (to_num(r[ci]) if ci is not None else 0.0)
                         for label, ci in qlick_col_by_label.items()},
    }
    venta_interna_rows.append(rec)

# Normalizar Gran Familia: uppercase + aplicar mapping de display names
for rec in venta_interna_rows:
    rec["gran_familia"] = display_family(rec["gran_familia"].upper())

# Conjunto de familias para el filtro (unión de las que tienen datos)
families = set()
for p in iqvia_products:
    families.add(p["familia"])
for rec in venta_interna_rows:
    families.add(rec["gran_familia"])
families = sorted(f for f in families if f)

# ---------------------------------------------------------------------------
# 5. Sanity checks por consola
# ---------------------------------------------------------------------------
print("→ Sanity checks:")
for (y, m) in window_default:
    print(f"  {yyyymm_label_es(y, m)}: Σ Siegfried-SI = {si_totals_by_month[(y, m)]:,.0f}")

last_label = month_labels[-1]
total_si_last     = sum(sum(pres["si"][last_label] for pres in p["presentaciones"])
                        for p in iqvia_products)
total_market_last = sum(p["mercado_molecula"][last_label] for p in iqvia_products)
ms_global = (total_si_last / total_market_last * 100) if total_market_last else 0
print(f"  MS% global Siegfried (último mes {last_label}, moléculas Siegfried): {ms_global:.2f}%")

fam_vi = sorted({rec['gran_familia'] for rec in venta_interna_rows})
print(f"  Familias en Venta Interna: {len(fam_vi)} → {fam_vi}")

# ---------------------------------------------------------------------------
# 6. Render HTML
# ---------------------------------------------------------------------------
default_window_labels = [yyyymm_label_es(y, m) for (y, m) in window_default]

# Metadata para la cabecera: última actualización + meses más recientes con datos
last_update = datetime.now().strftime("%d/%m/%Y %H:%M")

# Último mes IQVIA con Siegfried-SI > 0
latest_iqvia = month_labels[-1] if month_labels else "—"

# Último mes QLICK con alguna unidad > 0 (puede no coincidir con IQVIA)
latest_qlick = "—"
for label in reversed(month_labels):
    total = sum(rec["data"].get(label, 0) for rec in venta_interna_rows)
    if total > 0:
        latest_qlick = label
        break

print(f"  última actualización: {last_update}")
print(f"  último mes IQVIA con datos: {latest_iqvia}")
print(f"  último mes Venta Interna con datos: {latest_qlick}")

payload = {
    "all_months":     month_labels,
    "default_months": default_window_labels,
    "families":       families,
    "iqvia_products": iqvia_products,
    "venta_interna":  venta_interna_rows,
    "meta": {
        "last_update":  last_update,
        "latest_iqvia": latest_iqvia,
        "latest_qlick": latest_qlick,
    },
}

HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Dashboard Seguimiento Lanzamientos - Siegfried</title>
<style>
  :root {
    --bg: #f6f7fb; --card: #fff; --ink: #1f2937; --muted: #6b7280;
    --accent: #1e3a8a; --accent2: #2563eb; --border: #e5e7eb;
    --sig: #1e3a8a; --mercado: #6b7280; --ms: #047857;
  }
  * { box-sizing: border-box; }
  body { margin: 0; font-family: -apple-system, "Segoe UI", Roboto, Arial, sans-serif;
         background: var(--bg); color: var(--ink); }
  header { background: var(--accent); color: #fff; padding: 18px 28px; }
  header h1 { margin: 0; font-size: 20px; font-weight: 600; letter-spacing: 0.2px; }
  header .sub { opacity: 0.85; font-size: 13px; margin-top: 4px; }
  .container { max-width: 1500px; margin: 0 auto; padding: 18px 28px 40px; }
  .toolbar { display: flex; flex-wrap: wrap; gap: 14px; align-items: center;
             background: var(--card); padding: 14px 18px; border-radius: 10px;
             box-shadow: 0 1px 2px rgba(0,0,0,0.04); margin-bottom: 18px;
             border: 1px solid var(--border); }
  .toolbar label { font-size: 13px; color: var(--muted); margin-right: 6px; }
  .toolbar select { padding: 6px 10px; border-radius: 6px; border: 1px solid var(--border);
                    font-size: 14px; background: #fff; }
  .seg-toggle { display: inline-flex; border: 1px solid var(--border); border-radius: 6px;
                overflow: hidden; }
  .seg-toggle button { background: #fff; border: 0; padding: 7px 14px; cursor: pointer;
                       font-size: 13px; color: var(--ink); }
  .seg-toggle button.active { background: var(--accent2); color: #fff; }
  .months-card { background: var(--card); border: 1px solid var(--border);
                 border-radius: 10px; padding: 14px 18px; margin-bottom: 18px;
                 box-shadow: 0 1px 2px rgba(0,0,0,0.03); }
  .months-card .row1 { display: flex; flex-wrap: wrap; gap: 10px; align-items: center;
                       margin-bottom: 10px; }
  .months-card .label { font-size: 13px; color: var(--muted); font-weight: 500; }
  .preset-btns { display: inline-flex; gap: 6px; flex-wrap: wrap; }
  .preset-btns button { background: #fff; border: 1px solid var(--border);
                        padding: 5px 10px; border-radius: 6px; font-size: 12px;
                        cursor: pointer; color: var(--ink); }
  .preset-btns button:hover { background: #f3f4f6; }
  .chip-list { display: flex; flex-wrap: wrap; gap: 6px; }
  .chip { background: #fff; border: 1px solid var(--border); padding: 5px 10px;
          border-radius: 16px; font-size: 12px; cursor: pointer; user-select: none;
          color: var(--muted); transition: all 0.1s; }
  .chip:hover { border-color: var(--accent2); }
  .chip.selected { background: var(--accent2); color: #fff; border-color: var(--accent2); }
  /* Resalte del año más reciente */
  .chip.latest-year { border-color: var(--accent); border-width: 2px;
                      color: var(--accent); font-weight: 600; padding: 4px 10px; }
  .chip.latest-year:hover { background: #eef2ff; }
  .chip.latest-year.selected { background: var(--accent); border-color: var(--accent);
                               color: #fff; }
  .chip-count { font-size: 11px; color: var(--muted); margin-left: 6px; }
  .year-row { display: flex; flex-wrap: wrap; gap: 6px; align-items: center;
              margin: 4px 0 10px; }
  .year-row .label-inline { font-size: 12px; color: var(--muted); margin-right: 4px; }
  .year-btn { background: #fff; border: 1px solid var(--border); padding: 5px 10px;
              border-radius: 6px; font-size: 12px; cursor: pointer; color: var(--ink); }
  .year-btn:hover { background: #f3f4f6; }
  .year-btn.open { background: #eef2ff; border-color: var(--accent2); color: var(--accent); }
  .year-chips-section { background: #f9fafb; border: 1px dashed var(--border);
                        border-radius: 6px; padding: 8px 10px; margin: 4px 0 10px; }
  .year-chips-section .section-label { font-size: 11px; color: var(--muted);
                                       margin-bottom: 6px; text-transform: uppercase;
                                       letter-spacing: 0.5px; }
  .recent-label { font-size: 11px; color: var(--muted); margin: 8px 0 6px;
                  text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600; }
  section.card { background: var(--card); border: 1px solid var(--border);
                 border-radius: 10px; padding: 18px 20px; margin-bottom: 22px;
                 box-shadow: 0 1px 2px rgba(0,0,0,0.03); }
  section.card h2 { margin: 0 0 18px; font-size: 28px; color: var(--accent);
                    display: flex; justify-content: space-between; align-items: baseline;
                    font-weight: 800; letter-spacing: -0.3px;
                    border-bottom: 3px solid var(--accent); padding-bottom: 10px; }
  section.card h2 .sub-h { font-size: 12px; font-weight: 400; color: var(--muted);
                           text-transform: none; letter-spacing: normal; }
  table { width: 100%; border-collapse: separate; border-spacing: 0; font-size: 13px; }
  th, td { padding: 7px 10px; text-align: right; border-bottom: 1px solid var(--border); }
  th:first-child, td:first-child { text-align: left; }
  th { background: #f9fafb; font-weight: 600; color: var(--muted); font-size: 12px;
       text-transform: uppercase; letter-spacing: 0.4px;
       position: sticky; top: 0; z-index: 20;
       box-shadow: 0 2px 0 var(--border); }
  tr.entity-header td { background: #eef2ff; font-weight: 600; color: var(--accent);
                        border-top: 2px solid var(--accent); }
  tr.prod-header td { background: #eef2ff; color: var(--accent);
                      border-top: 3px solid var(--accent); padding-top: 10px; padding-bottom: 10px; }
  tr.prod-header .prod-name { font-weight: 700; font-size: 15px; }
  tr.prod-header .prod-tag  { font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px;
                              color: var(--muted); margin-left: 10px; }
  tr.prod-header .prod-sub  { font-weight: 500; color: var(--ink); font-size: 13px; }
  tr.prod-header .prod-fam  { color: var(--muted); font-size: 11px; margin-left: 8px; }
  tr.prod-header td.prod-total { font-weight: 700; color: var(--accent); font-size: 14px; }
  tr.fila-ms-total td { background: #ecfdf5; font-weight: 700; }
  tr.fila-ms-total td:first-child { color: var(--ms); padding-left: 22px; }
  tr.fila-ms-total td:not(:first-child) { color: var(--ms); font-size: 14px; }
  tr.pres-header td { background: #f9fafb; font-weight: 600; color: var(--ink);
                      padding-left: 14px; border-top: 1px solid var(--border); }
  tr.fila-si  td:first-child { color: var(--sig); font-weight: 500; padding-left: 32px; }
  tr.fila-si  td:not(:first-child) { color: var(--sig); }
  tr.fila-mdo td:first-child { color: var(--mercado); padding-left: 22px; font-weight: 500; }
  tr.fila-mdo td { background: #fafafa; }
  tr.fila-ms  td { font-weight: 600; }
  tr.fila-ms  td:first-child { color: var(--ms); padding-left: 32px; }
  tr.fila-ms  td:not(:first-child) { color: var(--ms); }
  tr.fila-evol td:first-child { color: #92400e; font-style: italic; padding-left: 32px;
                                font-size: 11px; font-weight: 500; }
  tr.fila-evol td:not(:first-child) { font-size: 11px; font-weight: 600; }
  tr.fila-evol .up   { color: #047857; }
  tr.fila-evol .down { color: #b91c1c; }
  tr.fila-evol .flat { color: var(--muted); }
  .num { font-variant-numeric: tabular-nums; }
  tr.familia-header td { background: #f3f4f6; font-weight: 700; color: var(--accent);
                         border-top: 2px solid var(--accent); padding-left: 10px; }
  .empty { color: var(--muted); padding: 18px; text-align: center; font-style: italic; }
  .footer { color: var(--muted); font-size: 12px; margin-top: 8px; }
</style>
</head>
<body>

<header>
  <h1>Dashboard Seguimiento Lanzamientos - Siegfried</h1>
  <div class="sub">
    <span id="hdr-meta"></span>
  </div>
  <div class="sub">Ventana mostrada: <span id="hdr-meses"></span></div>
</header>

<div class="container">

  <div class="toolbar">
    <label for="filtro-familia">Familia:</label>
    <select id="filtro-familia">
      <option value="__all__">Todas</option>
    </select>
    <span style="flex:1"></span>
    <span style="font-size:13px;color:var(--muted);">Vista IQVIA:</span>
    <div class="seg-toggle" id="iqvia-toggle">
      <button data-view="molecula" class="active">Por Molécula</button>
      <button data-view="atc">Por ATC</button>
    </div>
    <span style="font-size:13px;color:var(--muted);">Nivel:</span>
    <div class="seg-toggle" id="agg-toggle">
      <button data-agg="familia" class="active">Sólo Familia</button>
      <button data-agg="presentacion">Con Presentaciones</button>
    </div>
    <span style="font-size:13px;color:var(--muted);">Evolución:</span>
    <div class="seg-toggle" id="evol-toggle">
      <button data-evol="off" class="active">Ocultar</button>
      <button data-evol="on">Mostrar</button>
    </div>
  </div>

  <div class="months-card">
    <div class="row1">
      <span class="label">Meses a mostrar <span class="chip-count" id="months-count"></span></span>
      <div class="preset-btns" id="month-presets">
        <button data-n="3">Últimos 3</button>
        <button data-n="6">Últimos 6</button>
        <button data-n="12">Últimos 12</button>
        <button data-n="all">Todos</button>
        <button data-n="none">Ninguno</button>
      </div>
    </div>
    <div class="chip-list" id="months-chips"></div>
  </div>

  <section class="card">
    <h2>IQVIA <span class="sub-h">Pack Siegfried vs Mercado · MS% = Siegfried / Total Mercado</span></h2>
    <div id="iqvia-table-wrap"></div>
  </section>

  <section class="card">
    <h2>VENTA INTERNA <span class="sub-h">Packs por Familia / Presentación</span></h2>
    <div id="vi-table-wrap"></div>
  </section>

  <div class="footer">Fuente: IQUVIA_PM.xlsx · QLICK_VTA_INTERNA.xlsx · MAESTRO.xlsx</div>
</div>

<script>
const DATA = __PAYLOAD__;

const fmt = (n) => {
  if (n === null || n === undefined || isNaN(n)) return "—";
  if (n === 0) return "0";
  return Math.round(n).toLocaleString("es-AR");
};
const fmtPct = (n) => (n === 0 || !isFinite(n)) ? "0,0%" :
  n.toLocaleString("es-AR", {minimumFractionDigits: 1, maximumFractionDigits: 1}) + "%";

// --- Estado ---
let selectedMonths = DATA.default_months.slice();  // arranca con los 6 default
let currentView    = "molecula";   // molecula | atc
let aggLevel       = "familia";      // familia | presentacion
let showEvol       = false;          // mostrar fila Δ mes a mes

const fmtEvol = (cur, prev) => {
  if (prev == null || prev === 0) return `<span class="flat">—</span>`;
  if (cur == null) return `<span class="flat">—</span>`;
  const delta = (cur - prev) / prev * 100;
  if (Math.abs(delta) < 0.05) return `<span class="flat">0,0%</span>`;
  const arrow = delta > 0 ? "▲" : "▼";
  const klass = delta > 0 ? "up"  : "down";
  return `<span class="${klass}">${arrow} ${delta.toLocaleString("es-AR",{minimumFractionDigits:1,maximumFractionDigits:1})}%</span>`;
};

// --- Filtro familia ---
const sel = document.getElementById("filtro-familia");
DATA.families.forEach(f => {
  const o = document.createElement("option");
  o.value = f; o.textContent = f;
  sel.appendChild(o);
});
sel.addEventListener("change", renderAll);

// --- Toggle Molécula / ATC ---
document.querySelectorAll("#iqvia-toggle button").forEach(b => {
  b.addEventListener("click", () => {
    document.querySelectorAll("#iqvia-toggle button").forEach(x => x.classList.remove("active"));
    b.classList.add("active");
    currentView = b.dataset.view;
    renderIqvia();
  });
});

// --- Toggle Nivel: Presentación / Familia ---
document.querySelectorAll("#agg-toggle button").forEach(b => {
  b.addEventListener("click", () => {
    document.querySelectorAll("#agg-toggle button").forEach(x => x.classList.remove("active"));
    b.classList.add("active");
    aggLevel = b.dataset.agg;
    renderIqvia();
  });
});

// --- Toggle Evolución ---
document.querySelectorAll("#evol-toggle button").forEach(b => {
  b.addEventListener("click", () => {
    document.querySelectorAll("#evol-toggle button").forEach(x => x.classList.remove("active"));
    b.classList.add("active");
    showEvol = b.dataset.evol === "on";
    renderIqvia();
    renderVI();
  });
});

// --- Chips de meses: últimos 12 + colapsables por año ---
const RECENT_N = 12;
const recentMonths = DATA.all_months.slice(-RECENT_N);
const olderMonths  = DATA.all_months.slice(0, -RECENT_N);
const monthsByYear = {};
olderMonths.forEach(m => {
  const y = m.split('-')[1];
  if (!monthsByYear[y]) monthsByYear[y] = [];
  monthsByYear[y].push(m);
});
const olderYears = Object.keys(monthsByYear).sort();
const expandedYears = new Set();

const LATEST_YEAR = DATA.all_months[DATA.all_months.length - 1].split('-')[1];
function chipHTML(m) {
  const cls = ["chip"];
  if (selectedMonths.includes(m)) cls.push("selected");
  if (m.split('-')[1] === LATEST_YEAR) cls.push("latest-year");
  return `<span class="${cls.join(' ')}" data-month="${m}">${m}</span>`;
}

function renderChips() {
  const wrap = document.getElementById("months-chips");
  let html = "";
  if (olderYears.length > 0) {
    html += `<div class="year-row"><span class="label-inline">Histórico:</span>`;
    olderYears.forEach(y => {
      const open = expandedYears.has(y);
      html += `<button class="year-btn${open ? " open" : ""}" data-year="${y}">${open ? "▼" : "▶"} ${y}</button>`;
    });
    html += `</div>`;
    olderYears.forEach(y => {
      if (expandedYears.has(y)) {
        html += `<div class="year-chips-section">` +
          `<div class="section-label">Año ${y}</div>` +
          `<div class="chip-list">` +
            monthsByYear[y].map(chipHTML).join("") +
          `</div></div>`;
      }
    });
  }
  html += `<div class="recent-label">Últimos ${RECENT_N} meses</div>`;
  html += `<div class="chip-list">${recentMonths.map(chipHTML).join("")}</div>`;
  wrap.innerHTML = html;

  // Listeners
  wrap.querySelectorAll(".chip").forEach(c => {
    c.addEventListener("click", () => toggleMonth(c.dataset.month));
  });
  wrap.querySelectorAll(".year-btn").forEach(b => {
    b.addEventListener("click", () => {
      const y = b.dataset.year;
      if (expandedYears.has(y)) expandedYears.delete(y);
      else expandedYears.add(y);
      renderChips();
    });
  });
}

function toggleMonth(m) {
  if (selectedMonths.includes(m)) {
    selectedMonths = selectedMonths.filter(x => x !== m);
  } else {
    selectedMonths.push(m);
    selectedMonths.sort((a, b) => DATA.all_months.indexOf(a) - DATA.all_months.indexOf(b));
  }
  // refresca el estado visual de TODOS los chips visibles
  document.querySelectorAll("#months-chips .chip").forEach(c => {
    c.classList.toggle("selected", selectedMonths.includes(c.dataset.month));
  });
  renderAll();
}

renderChips();

// --- Atajos ---
document.querySelectorAll("#month-presets button").forEach(btn => {
  btn.addEventListener("click", () => {
    const n = btn.dataset.n;
    if (n === "all")        selectedMonths = DATA.all_months.slice();
    else if (n === "none")  selectedMonths = [];
    else                    selectedMonths = DATA.all_months.slice(-parseInt(n));
    document.querySelectorAll("#months-chips .chip").forEach(c => {
      c.classList.toggle("selected", selectedMonths.includes(c.dataset.month));
    });
    renderAll();
  });
});

function updateHeader() {
  document.getElementById("hdr-meses").textContent =
    selectedMonths.length ? selectedMonths.join(" · ") : "(sin meses seleccionados)";
  document.getElementById("months-count").textContent =
    `(${selectedMonths.length} de ${DATA.all_months.length})`;
}

// Cabecera de metadata (una sola vez al cargar)
(function setMeta() {
  const m = DATA.meta || {};
  document.getElementById("hdr-meta").innerHTML =
    `Última actualización: <b>${m.last_update || "—"}</b>` +
    ` &middot; Datos IQVIA hasta: <b>${m.latest_iqvia || "—"}</b>` +
    ` &middot; Venta Interna hasta: <b>${m.latest_qlick || "—"}</b>`;
})();

// Construye entidades a nivel FAMILIA. Cada entidad agrupa productos por familia.
// { familia, productos: [iqvia_products], mercado_mol, mercado_atc }
function buildFamilyEntities() {
  const filtro = sel.value;
  const filtered = (filtro === "__all__")
      ? DATA.iqvia_products
      : DATA.iqvia_products.filter(p => p.familia === filtro);

  const fams = {};
  filtered.forEach(p => {
    if (!fams[p.familia]) {
      fams[p.familia] = { familia: p.familia, productos: [],
                          mercado_mol: {}, mercado_atc: {} };
      DATA.all_months.forEach(m => {
        fams[p.familia].mercado_mol[m] = 0;
        fams[p.familia].mercado_atc[m] = 0;
      });
    }
    fams[p.familia].productos.push(p);
    DATA.all_months.forEach(m => {
      fams[p.familia].mercado_mol[m] += (p.mercado_molecula[m] || 0);
      fams[p.familia].mercado_atc[m] += (p.mercado_atc[m]      || 0);
    });
  });
  return Object.values(fams).sort((a, b) => a.familia.localeCompare(b.familia));
}

function renderIqvia() {
  const entities = buildFamilyEntities();
  const wrap = document.getElementById("iqvia-table-wrap");
  if (entities.length === 0 || selectedMonths.length === 0) {
    wrap.innerHTML = `<div class="empty">Sin datos para esta selección.</div>`;
    return;
  }
  const viewIsMol = currentView === "molecula";
  const months    = selectedMonths;
  const ths       = months.map(m => `<th>${m}</th>`).join("");
  let html = `<table><thead><tr><th style="min-width:280px">Familia / Producto / Presentación</th>${ths}</tr></thead><tbody>`;

  entities.forEach(fam => {
    const mercado  = viewIsMol ? fam.mercado_mol : fam.mercado_atc;
    const subLabel = viewIsMol ? "Molécula"      : "ATC";

    // Subtitle: si la familia tiene 1 producto -> molécula/ATC; si tiene N -> lista de productos
    let subtitle, tagLabel;
    if (fam.productos.length === 1) {
      const p = fam.productos[0];
      subtitle = viewIsMol ? p.molecula : p.atc;
      tagLabel = subLabel;
    } else {
      subtitle = fam.productos.map(p => p.producto).join(" + ");
      tagLabel = "Productos";
    }

    // Pack Siegfried total de la familia
    const famPack = {};
    months.forEach(m => {
      famPack[m] = fam.productos.reduce(
        (sum, p) => sum + p.presentaciones.reduce((s, pr) => s + (pr.si[m] || 0), 0), 0);
    });

    // === BLUE HEADER: nombre familia + MERCADO TOTAL por mes ===
    html += `<tr class="prod-header">` +
      `<td><span class="prod-name">${fam.familia}</span>` +
      ` <span class="prod-tag">${tagLabel}:</span> <span class="prod-sub">${subtitle}</span>` +
      `</td>` +
      months.map(m => `<td class="num prod-total">${fmt(mercado[m])}</td>`).join("") +
      `</tr>`;
    if (showEvol) {
      html += `<tr class="fila-evol"><td>Δ Mercado vs mes ant.</td>` +
        months.map((m, i) => `<td class="num">${i === 0 ? '<span class="flat">—</span>' : fmtEvol(mercado[m], mercado[months[i-1]])}</td>`).join("") + `</tr>`;
    }

    // === Total {Familia}: suma de Packs Siegfried ===
    html += `<tr class="fila-si"><td>Total ${fam.familia}</td>` +
      months.map(m => `<td class="num">${fmt(famPack[m])}</td>`).join("") + `</tr>`;
    if (showEvol) {
      html += `<tr class="fila-evol"><td>Δ Total ${fam.familia} vs mes ant.</td>` +
        months.map((m, i) => `<td class="num">${i === 0 ? '<span class="flat">—</span>' : fmtEvol(famPack[m], famPack[months[i-1]])}</td>`).join("") + `</tr>`;
    }

    // === MS % ===
    html += `<tr class="fila-ms-total"><td>MS %</td>` +
      months.map(m => {
        const ms = mercado[m] > 0 ? (famPack[m] / mercado[m] * 100) : 0;
        return `<td class="num">${fmtPct(ms)}</td>`;
      }).join("") + `</tr>`;

    // === EXPANDIR PRESENTACIONES (sólo si aggLevel === "presentacion") ===
    if (aggLevel === "presentacion") {
      fam.productos.forEach(p => {
        const pMercado = viewIsMol ? p.mercado_molecula : p.mercado_atc;
        p.presentaciones.forEach(pres => {
          const presLabel = fam.productos.length > 1
              ? `${p.producto} — ${pres.label}` : pres.label;
          html += `<tr class="pres-header"><td>${presLabel}</td>` +
            months.map(() => '<td></td>').join("") + `</tr>`;
          html += `<tr class="fila-si"><td>Pack Siegfried</td>` +
            months.map(m => `<td class="num">${fmt(pres.si[m])}</td>`).join("") + `</tr>`;
          if (showEvol) {
            html += `<tr class="fila-evol"><td>Δ Pack vs mes ant.</td>` +
              months.map((m, i) => `<td class="num">${i === 0 ? '<span class="flat">—</span>' : fmtEvol(pres.si[m], pres.si[months[i-1]])}</td>`).join("") + `</tr>`;
          }
          html += `<tr class="fila-ms"><td>MS %</td>` +
            months.map(m => {
              const denom = pMercado[m] || 0;
              const ms = denom > 0 ? (pres.si[m] / denom * 100) : 0;
              return `<td class="num">${fmtPct(ms)}</td>`;
            }).join("") + `</tr>`;
        });
      });
    }
  });
  html += `</tbody></table>`;
  wrap.innerHTML = html;
}

function renderVI() {
  const filtro = sel.value;
  const rows = (filtro === "__all__") ? DATA.venta_interna
                                       : DATA.venta_interna.filter(r => r.gran_familia === filtro);
  const wrap = document.getElementById("vi-table-wrap");
  if (rows.length === 0 || selectedMonths.length === 0) {
    wrap.innerHTML = `<div class="empty">Sin datos para esta selección.</div>`;
    return;
  }
  const groups = {};
  rows.forEach(r => {
    if (!groups[r.gran_familia]) groups[r.gran_familia] = [];
    groups[r.gran_familia].push(r);
  });
  const months = selectedMonths;
  const ths = months.map(m => `<th>${m}</th>`).join("");
  let html = `<table><thead><tr><th style="min-width:280px">Familia / Presentación</th>${ths}</tr></thead><tbody>`;
  Object.keys(groups).sort().forEach(fam => {
    const totals = months.map(m =>
      groups[fam].reduce((acc, r) => acc + (r.data[m] || 0), 0));
    html += `<tr class="familia-header"><td>${fam}</td>` +
      totals.map(v => `<td class="num">${fmt(v)}</td>`).join("") + `</tr>`;
    if (showEvol) {
      html += `<tr class="fila-evol"><td style="padding-left:14px">Δ Familia vs mes ant.</td>` +
        totals.map((v, i) => `<td class="num">${i === 0 ? '<span class="flat">—</span>' : fmtEvol(v, totals[i-1])}</td>`).join("") + `</tr>`;
    }
    groups[fam].forEach(r => {
      html += `<tr><td style="padding-left:22px;color:var(--muted);">${r.presentacion}</td>` +
        months.map(m => `<td class="num">${fmt(r.data[m])}</td>`).join("") + `</tr>`;
      if (showEvol) {
        html += `<tr class="fila-evol"><td>Δ vs mes ant.</td>` +
          months.map((m, i) => `<td class="num">${i === 0 ? '<span class="flat">—</span>' : fmtEvol(r.data[m], r.data[months[i-1]])}</td>`).join("") + `</tr>`;
      }
    });
  });
  html += `</tbody></table>`;
  wrap.innerHTML = html;
}

function renderAll() {
  updateHeader();
  renderIqvia();
  renderVI();
}

renderAll();
</script>
</body>
</html>
"""

html = HTML_TEMPLATE.replace("__PAYLOAD__", json.dumps(payload, ensure_ascii=False))
OUT_HTML.write_text(html, encoding="utf-8")
print(f"→ Dashboard generado: {OUT_HTML}")
print(f"   Tamaño: {OUT_HTML.stat().st_size:,} bytes")
